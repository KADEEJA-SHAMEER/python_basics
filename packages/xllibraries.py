# from openpyxl import Workbook
# wb=Workbook()
# ws=wb.active
# ws['A1'] = "Name"
# ws['B1'] = "Age"
#
# ws.append(["Tom", 25])
# ws.append(["Jerry", 22])
# wb.save("sample1.xlsx")

from openpyxl import load_workbook

# Load the existing workbook
wb = load_workbook("sample1.xlsx")
ws = wb.active


ws.append(["Spike", 30])


ws['C1'] = "City"

ws['C2'] = "New York"
ws['C3'] = "Los Angeles"
ws['C4'] = "London"
wb.save("sample1.xlsx")

